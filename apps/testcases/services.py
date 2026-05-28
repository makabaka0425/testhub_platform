from __future__ import annotations

import re
import uuid
from dataclasses import dataclass
from io import BytesIO
from typing import Dict, List, Optional

from django.core.files.base import ContentFile
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.users.models import User
from apps.versions.models import Version

from .models import TestCase, TestCaseImportRecord


TEMPLATE_VERSION = 'v1'
TEMPLATE_SHEET_NAME = '用例导入模板'
INSTRUCTION_SHEET_NAME = '填写说明'


@dataclass
class ImportSummary:
    total_rows: int
    success_count: int
    failed_count: int
    skip_count: int
    failure_details: List[dict]


class TestCaseImportTemplateService:
    HEADERS = [
        '用例标题*',
        '前置条件',
        '操作步骤*',
        '预期结果*',
        '优先级',
        '测试类型',
        '关联版本',
    ]

    SAMPLE_ROW = [
        '用户登录成功校验',
        '用户已注册且账号状态正常',
        '1. 打开登录页\n2. 输入正确用户名和密码\n3. 点击登录按钮',
        '页面跳转到首页，并显示登录成功信息',
        '高',
        '功能测试',
        'V1.0,V1.1',
    ]

    INSTRUCTIONS = [
        ['模板版本', TEMPLATE_VERSION],
        ['说明', '请勿修改表头名称，按模板字段填写并上传'],
        ['默认处理', '模板中未提供的字段将使用系统默认值：描述为空、状态=draft(草稿)、标签=[]、指派人为空'],
        ['用例标题*', '必填，测试用例标题，建议唯一且明确'],
        ['前置条件', '可选'],
        ['操作步骤*', '必填，支持多行文本'],
        ['预期结果*', '必填，支持多行文本'],
        ['优先级', '可选，支持中文或英文：低/中/高/紧急 或 low/medium/high/critical；默认 medium(中)'],
        ['测试类型', '可选，支持中文或英文：功能测试/集成测试/API测试/UI测试/性能测试/安全测试 或 functional/integration/api/ui/performance/security；默认 functional(功能测试)'],
        ['中英文示例', '优先级示例：高 或 high；测试类型示例：功能测试 或 functional'],
        ['关联版本', '可选，填写当前项目下的版本名称，多个版本用英文逗号分隔'],
        ['关联项目', '上传时由页面选择，不需要在 Excel 中填写'],
    ]

    @classmethod
    def build_template(cls) -> BytesIO:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = TEMPLATE_SHEET_NAME
        worksheet.append(cls.HEADERS)
        worksheet.append(cls.SAMPLE_ROW)

        header_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
        required_fill = PatternFill(fill_type='solid', fgColor='FDE2E2')
        header_font = Font(bold=True)

        for index, header in enumerate(cls.HEADERS, start=1):
            cell = worksheet.cell(row=1, column=index)
            cell.font = header_font
            cell.fill = required_fill if header.endswith('*') else header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row in worksheet.iter_rows(min_row=2, max_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        column_widths = [24, 24, 40, 34, 14, 18, 22]
        for idx, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[chr(64 + idx)].width = width

        instruction_sheet = workbook.create_sheet(INSTRUCTION_SHEET_NAME)
        for row in cls.INSTRUCTIONS:
            instruction_sheet.append(row)
        instruction_sheet.column_dimensions['A'].width = 22
        instruction_sheet.column_dimensions['B'].width = 80
        for row in instruction_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        for cell in instruction_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output


class TestCaseExcelImportService:
    PRIORITY_MAP = {
        'low': 'low', '低': 'low',
        'medium': 'medium', '中': 'medium',
        'high': 'high', '高': 'high',
        'critical': 'critical', '紧急': 'critical',
    }
    STATUS_MAP = {
        'draft': 'draft', '草稿': 'draft',
        'active': 'active', '激活': 'active',
        'deprecated': 'deprecated', '废弃': 'deprecated',
    }
    TEST_TYPE_MAP = {
        'functional': 'functional', '功能测试': 'functional',
        'integration': 'integration', '集成测试': 'integration',
        'api': 'api', 'API测试': 'api',
        'ui': 'ui', 'UI测试': 'ui',
        'performance': 'performance', '性能测试': 'performance',
        'security': 'security', '安全测试': 'security',
    }

    HEADER_ALIASES = {
        '用例标题*': 'title',
        '标题*': 'title',
        '前置条件': 'preconditions',
        '操作步骤*': 'steps',
        '预期结果*': 'expected_result',
        '优先级': 'priority',
        '测试类型': 'test_type',
        '关联版本': 'versions',
    }
    EXPECTED_HEADER_FIELDS = [
        'title',
        'preconditions',
        'steps',
        'expected_result',
        'priority',
        'test_type',
        'versions',
    ]

    @classmethod
    def import_record(cls, record: TestCaseImportRecord) -> ImportSummary:
        workbook = load_workbook(record.import_file.path)
        if TEMPLATE_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f'Excel 中缺少工作表: {TEMPLATE_SHEET_NAME}')

        worksheet = workbook[TEMPLATE_SHEET_NAME]
        raw_headers = [cls._normalize_header(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        effective_headers = [header for header in raw_headers if header]
        header_fields = [cls.HEADER_ALIASES.get(header) for header in effective_headers]

        if None in header_fields or header_fields != cls.EXPECTED_HEADER_FIELDS:
            raise ValueError('导入模板表头不匹配，请先下载最新模板后再导入')

        rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        effective_rows = [row for row in rows if any(cls._normalize_text(value) for value in row)]
        total_rows = len(effective_rows)
        success_count = 0
        failed_count = 0
        skip_count = 0
        failure_details: List[dict] = []

        version_map = {
            version.name: version
            for version in Version.objects.filter(projects=record.project).distinct()
        }

        for index, row in enumerate(effective_rows, start=2):
            row_values = list(row[:len(header_fields)])
            row_data = {
                field: cls._normalize_text(value)
                for field, value in zip(header_fields, row_values)
            }

            error_messages = cls._validate_row(row_data, version_map)
            if error_messages:
                failed_count += 1
                failure_details.append({
                    'row_number': index,
                    'title': row_data.get('title', ''),
                    'errors': error_messages,
                })
                continue

            version_names = cls._split_csv(row_data.get('versions'))

            with transaction.atomic():
                testcase = TestCase.objects.create(
                    project=record.project,
                    title=row_data['title'],
                    description='',
                    preconditions=row_data.get('preconditions', ''),
                    steps=row_data['steps'],
                    expected_result=row_data['expected_result'],
                    priority=cls.PRIORITY_MAP.get(row_data.get('priority', '').lower(), 'medium'),
                    status='draft',
                    test_type=cls.TEST_TYPE_MAP.get(row_data.get('test_type', '').lower(), 'functional'),
                    tags=[],
                    author=record.created_by,
                    assignee=None,
                )
                if version_names:
                    testcase.versions.set([version_map[name].id for name in version_names])

            success_count += 1
            record.progress = int(success_count / total_rows * 100) if total_rows else 100
            record.success_count = success_count
            record.failed_count = failed_count
            record.total_rows = total_rows
            record.save(update_fields=['progress', 'success_count', 'failed_count', 'total_rows', 'updated_at'])

        return ImportSummary(
            total_rows=total_rows,
            success_count=success_count,
            failed_count=failed_count,
            skip_count=skip_count,
            failure_details=failure_details,
        )

    @classmethod
    def build_failure_report(cls, record: TestCaseImportRecord) -> Optional[ContentFile]:
        if not record.failure_details:
            return None

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = '失败明细'
        worksheet.append(['行号', '标题', '失败原因'])
        for item in record.failure_details:
            worksheet.append([
                item.get('row_number'),
                item.get('title', ''),
                '；'.join(item.get('errors', [])),
            ])

        for column in ('A', 'B', 'C'):
            worksheet.column_dimensions[column].width = 20 if column != 'C' else 80

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return ContentFile(output.read(), name=f'{record.import_no}_failed_rows.xlsx')

    @classmethod
    def generate_import_no(cls) -> str:
        return f'IMP_{timezone.now().strftime("%Y%m%d%H%M%S")}_{uuid.uuid4().hex[:6].upper()}'

    @classmethod
    def _validate_row(cls, row_data: Dict[str, str], version_map: Dict[str, Version]) -> List[str]:
        errors = []
        if not row_data.get('title'):
            errors.append('标题不能为空')
        if not row_data.get('steps'):
            errors.append('操作步骤不能为空')
        if not row_data.get('expected_result'):
            errors.append('预期结果不能为空')

        priority = row_data.get('priority', '')
        if priority and priority.lower() not in cls.PRIORITY_MAP and priority not in cls.PRIORITY_MAP:
            errors.append(f'优先级无效: {priority}')

        test_type = row_data.get('test_type', '')
        if test_type and test_type.lower() not in cls.TEST_TYPE_MAP and test_type not in cls.TEST_TYPE_MAP:
            errors.append(f'测试类型无效: {test_type}')

        version_names = cls._split_csv(row_data.get('versions'))
        invalid_versions = [name for name in version_names if name not in version_map]
        if invalid_versions:
            errors.append(f'版本不存在: {", ".join(invalid_versions)}')

        return errors

    @staticmethod
    def _get_assignee(username: str) -> Optional[User]:
        if not username:
            return None
        return User.objects.filter(username=username).first()

    @staticmethod
    def _split_csv(text: str) -> List[str]:
        if not text:
            return []
        return [item.strip() for item in str(text).split(',') if item and item.strip()]

    @staticmethod
    def _normalize_text(value) -> str:
        if value is None:
            return ''
        return str(value).strip()

    @staticmethod
    def _normalize_header(value) -> str:
        text = TestCaseExcelImportService._normalize_text(value)
        if not text:
            return ''
        text = text.replace('\ufeff', '').replace('\u200b', '')
        return re.sub(r'\s+', '', text)
